import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_exception_handling_excel():
    # Data for the exception handling plan
    data = {
        'Exception Category': [
            'Database Connection',
            'Model Not Found', 
            'Data Integrity',
            'File Upload Processing',
            'Authentication',
            'HTTP Request Errors',
            'Template Rendering',
            'Configuration Issues',
            'Data Validation',
            'File System Operations',
            'CSRF Protection',
            'Session Management',
            'JSON Processing',
            'Date/Time Processing',
            'Memory/Performance',
            'Network Connectivity',
            'Form Validation',
            'Database Transaction',
            'Static Files',
            'Email Operations'
        ],
        'Possible Exceptions': [
            'django.db.utils.OperationalError, psycopg2.OperationalError, django.db.utils.InterfaceError',
            'UserProfile.DoesNotExist, Organization.DoesNotExist, BSCEntry.DoesNotExist',
            'django.db.utils.IntegrityError, django.core.exceptions.ValidationError',
            'pandas.errors.EmptyDataError, pandas.errors.ParserError, UnicodeDecodeError',
            'django.contrib.auth.AuthenticationFailed, django.contrib.auth.PermissionDenied',
            'django.http.Http404, django.http.Http500, django.core.exceptions.SuspiciousOperation',
            'django.template.TemplateDoesNotExist, django.template.TemplateSyntaxError',
            'django.core.exceptions.ImproperlyConfigured, ImportError, KeyError',
            'ValueError, TypeError, KeyError',
            'FileNotFoundError, PermissionError, OSError',
            'django.middleware.csrf.CsrfViewMiddleware, django.core.exceptions.SuspiciousOperation',
            'django.contrib.sessions.exceptions.SessionInterrupted, django.contrib.sessions.exceptions.InvalidSessionKey',
            'json.JSONDecodeError, json.JSONEncodeError',
            'ValueError (invalid dates), TypeError (wrong date format)',
            'MemoryError, TimeoutError',
            'requests.exceptions.ConnectionError, requests.exceptions.Timeout',
            'django.forms.ValidationError, django.core.exceptions.ValidationError',
            'django.db.transaction.TransactionManagementError, django.db.utils.DatabaseError',
            'django.contrib.staticfiles.exceptions.IncorrectStaticFilesBackend, FileNotFoundError',
            'django.core.mail.BadHeaderError, smtplib.SMTPException'
        ],
        'Handling Strategy': [
            'try:\n    # Database operation\nexcept OperationalError:\n    logging.error("DB connection failed")\n    # Implement retry logic or fallback',
            'try:\n    profile = user.userprofile\nexcept UserProfile.DoesNotExist:\n    profile = None\n    # Create default profile or handle gracefully',
            'try:\n    BSCEntry.objects.create(...)\nexcept IntegrityError:\n    messages.error("Duplicate entry or constraint violation")\nexcept ValidationError as e:\n    messages.error(str(e))',
            'try:\n    if file.name.endswith(".csv"):\n        df = pd.read_csv(file, encoding="utf-8")\n    else:\n        df = pd.read_excel(file)\nexcept EmptyDataError:\n    raise ValidationError("Uploaded file is empty")\nexcept ParserError:\n    raise ValidationError("Invalid file format or corrupted file")\nexcept UnicodeDecodeError:\n    raise ValidationError("File encoding issue. Please use UTF-8")',
            'try:\n    user = authenticate(username, password)\n    if not user:\n        raise AuthenticationFailed\nexcept AuthenticationFailed:\n    messages.error("Invalid username or password")\nexcept PermissionDenied:\n    messages.error("You do not have permission to access this resource")',
            'try:\n    return view_function(request)\nexcept Http404:\n    return JsonResponse({"error": "Resource not found"}, status=404)\nexcept Http500:\n    return JsonResponse({"error": "Internal server error"}, status=500)\nexcept SuspiciousOperation:\n    return JsonResponse({"error": "Security violation detected"}, status=400)',
            'try:\n    return render(request, template_name, context)\nexcept TemplateDoesNotExist:\n    return render(request, "error/404.html", status=404)\nexcept TemplateSyntaxError:\n    logging.error("Template syntax error")\n    return render(request, "error/500.html", status=500)',
            'try:\n    validate_settings()\nexcept ImproperlyConfigured as e:\n    logging.critical(f"Configuration error: {e}")\n    # Exit gracefully or use defaults\nexcept ImportError as e:\n    logging.error(f"Missing dependency: {e}")\n    # Install missing package or use alternative',
            'try:\n    validate_bsc_data(data)\n    # Check required fields, data types, ranges\nexcept (ValueError, TypeError) as e:\n    messages.error(f"Invalid data format: {e}")\nexcept KeyError as e:\n    messages.error(f"Missing required field: {e}")',
            'try:\n    with open(file_path, "r") as f:\n        data = f.read()\nexcept FileNotFoundError:\n    messages.error("File not found")\nexcept PermissionError:\n    messages.error("Access denied to file")\nexcept OSError as e:\n    messages.error(f"File system error: {e}")',
            'try:\n    # Process form with CSRF token\nexcept SuspiciousOperation:\n    messages.error("Security violation: Invalid CSRF token")\n    return redirect("login")',
            'try:\n    request.session["key"] = value\nexcept SessionInterrupted:\n    # Redirect to login page\n    return redirect("login")\nexcept InvalidSessionKey:\n    # Clear session and redirect\n    request.session.flush()\n    return redirect("login")',
            'try:\n    data = json.loads(request.body)\nexcept JSONDecodeError:\n    return JsonResponse({"error": "Invalid JSON format"}, status=400)\nexcept JSONEncodeError:\n    return JsonResponse({"error": "Error encoding response"}, status=500)',
            'try:\n    date = datetime.strptime(date_str, "%Y-%m-%d")\nexcept ValueError:\n    messages.error("Invalid date format. Use YYYY-MM-DD")\nexcept TypeError:\n    messages.error("Date field must be a string")',
            'try:\n    # Process large file or dataset\nexcept MemoryError:\n    messages.error("File too large to process. Please use smaller file")\nexcept TimeoutError:\n    messages.error("Operation timed out. Please try again")',
            'try:\n    response = requests.get(url, timeout=30)\nexcept ConnectionError:\n    logging.error("Network connection failed")\n    # Use cached data or show offline message\nexcept Timeout:\n    logging.error("Request timed out")\n    # Retry or show timeout message',
            'try:\n    form = BSCForm(request.POST)\n    if form.is_valid():\n        form.save()\n    else:\n        for field, errors in form.errors.items():\n            messages.error(f"{field}: {errors}")\nexcept ValidationError as e:\n    messages.error(str(e))',
            'try:\n    with transaction.atomic():\n        # Database operations\n        pass\nexcept TransactionManagementError:\n    logging.error("Transaction management error")\n    # Rollback and retry\nexcept DatabaseError:\n    logging.error("Database operation failed")\n    # Handle gracefully',
            'try:\n    # Serve static files\nexcept IncorrectStaticFilesBackend:\n    logging.error("Static files backend configuration error")\nexcept FileNotFoundError:\n    logging.error("Static file not found")\n    # Use default or fallback',
            'try:\n    send_mail(subject, message, from_email, recipient_list)\nexcept BadHeaderError:\n    logging.error("Invalid email header")\nexcept SMTPException:\n    logging.error("Email sending failed")\n    # Queue for retry or notify admin'
        ],
        'Priority': [
            'High', 'High', 'High', 'High', 'High', 
            'Medium', 'Medium', 'Low', 'Medium', 'Medium',
            'High', 'Medium', 'Medium', 'Medium', 'Low',
            'Low', 'Medium', 'High', 'Low', 'Low'
        ],
        'Status': [
            'Not Implemented', 'Partially Implemented', 'Not Implemented', 'Partially Implemented', 'Partially Implemented',
            'Not Implemented', 'Not Implemented', 'Not Implemented', 'Not Implemented', 'Not Implemented',
            'Not Implemented', 'Not Implemented', 'Not Implemented', 'Not Implemented', 'Not Implemented',
            'Not Implemented', 'Not Implemented', 'Not Implemented', 'Not Implemented', 'Not Implemented'
        ]
    }
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Create Excel file
    filename = 'BSC_Exception_Handling_Plan.xlsx'
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Exception Handling Plan', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Exception Handling Plan']
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        high_priority_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        medium_priority_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        low_priority_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Style headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Style priority columns
        for row in range(2, len(df) + 2):
            priority_cell = worksheet[f'D{row}']
            if priority_cell.value == 'High':
                priority_cell.fill = high_priority_fill
            elif priority_cell.value == 'Medium':
                priority_cell.fill = medium_priority_fill
            elif priority_cell.value == 'Low':
                priority_cell.fill = low_priority_fill
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 50
        worksheet.column_dimensions['C'].width = 60
        worksheet.column_dimensions['D'].width = 10
        worksheet.column_dimensions['E'].width = 20
        
        # Wrap text for handling strategy column
        for row in range(2, len(df) + 2):
            worksheet[f'C{row}'].alignment = Alignment(wrap_text=True, vertical="top")
    
    print(f"Excel file '{filename}' created successfully!")
    print(f"File location: {os.path.abspath(filename)}")
    
    return filename

if __name__ == "__main__":
    try:
        create_exception_handling_excel()
    except ImportError:
        print("Please install required packages:")
        print("pip install pandas openpyxl")
    except Exception as e:
        print(f"Error creating Excel file: {e}") 